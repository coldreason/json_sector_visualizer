ROAD_CELL_WIDTH = 2
WIDE_DISTANCE_WIDTH = 1
NARROW_DISTANCE_WIDTH = 0

SEEDFILEPATH = "seed.xlsx"
OUTPUTFILEPATH = "res.xlsx"

class section:
    def __init__(self,row_distance_list,column_distance_list,dataShape,data):
        self.column_distance_list = column_distance_list
        self.row_distance_list = row_distance_list
        self.dataShape = dataShape
        self.data = data
        
class warehouse:
    def __init__(self,layout,sectionList):
        self.layout = layout
        self.sectionList = sectionList
        
def createDummyIntMap(warehouseList):
    buf = set()
    for warehouse in warehouseList:
        for section in warehouse.sectionList:
            
            for level3 in section.data:
                for level2 in level3:
                    buf = buf.union(set(level2))
    idToIntMap = {}
    idx = 1
    for i in list(buf):
        idToIntMap[str(i)] = idx
        idx= idx+1
    return idToIntMap

def createIntdymmyMap(idToIntMap):
    ret = {}
    for i in idToIntMap.keys():
        ret[idToIntMap[i]] = i
    return ret

def dummyToColor(dummyToIntMap):
    resMap = {}
    for i in dummyToIntMap.keys():
        resMap[str(dummyToIntMap[i])] = str(hex(random.randrange(1,16777215))[2:]).zfill(6)
    return resMap

def intToRom(numb):
    fir = ''
    if int(numb / 26) > 0:
        fir =  chr(64 + int(numb / 26))
    sec = chr(65 + numb % 26)
    return str(fir) + str(sec)

def getLargeMap(warehouse):
    width = 0
    height = 0
    if warehouse.layout == "I":
        for section in warehouse.sectionList:
            width = width + section.column_distance_list.count(True) *WIDE_DISTANCE_WIDTH + section.column_distance_list.count(False) *NARROW_DISTANCE_WIDTH + section.dataShape[0]
            height = max(height,(section.row_distance_list.count(True)*WIDE_DISTANCE_WIDTH + section.row_distance_list.count(False) *NARROW_DISTANCE_WIDTH + section.dataShape[1] * section.dataShape[2]))
        width = width + ROAD_CELL_WIDTH
    if warehouse.layout == "H":
        width = max(warehouse.sectionList[1].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH+warehouse.sectionList[1].column_distance_list.count(False) *NARROW_DISTANCE_WIDTH + warehouse.sectionList[1].dataShape[0],warehouse.sectionList[2].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[2].column_distance_list.count(False) * NARROW_DISTANCE_WIDTH + warehouse.sectionList[2].dataShape[0],)
        height = max((warehouse.sectionList[0].dataShape[1] * warehouse.sectionList[0].dataShape[2] + warehouse.sectionList[0].row_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[0].row_distance_list.count(False) * NARROW_DISTANCE_WIDTH),(warehouse.sectionList[3].dataShape[1] * warehouse.sectionList[3].dataShape[2] + warehouse.sectionList[3].row_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[3].row_distance_list.count(False) * NARROW_DISTANCE_WIDTH))
        height = max(height,ROAD_CELL_WIDTH+warehouse.sectionList[1].dataShape[1] * warehouse.sectionList[1].dataShape[2]+warehouse.sectionList[1].row_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[1].row_distance_list.count(False) * NARROW_DISTANCE_WIDTH +warehouse.sectionList[2].dataShape[1] * warehouse.sectionList[2].dataShape[2]+warehouse.sectionList[2].row_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[2].row_distance_list.count(False) * NARROW_DISTANCE_WIDTH)
        
        width = width + warehouse.sectionList[0].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[0].column_distance_list.count(False) * NARROW_DISTANCE_WIDTH + warehouse.sectionList[0].dataShape[0]
        width = width + warehouse.sectionList[3].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[3].column_distance_list.count(False) * NARROW_DISTANCE_WIDTH + warehouse.sectionList[3].dataShape[0]
        width = width + ROAD_CELL_WIDTH * 2
    return np.zeros([height,width],dtype=int)

def drawSection(section,offset,drawingMap):
    idx = offset.copy()
    for i in range(0,section.dataShape[1]):
        for j in range(0,section.dataShape[0]):
            for k in range(0,section.dataShape[2]):
                y = idx[0]+section.dataShape[2] * i + k
                x = idx[1] + j
                if k == 0:
                    drawingMap[y][x] = dummyToIntMap[str(section.data[j][i][k])] + 1000
                elif k == section.dataShape[2] - 1:
                    drawingMap[y][x] = dummyToIntMap[str(section.data[j][i][k])] + 2000
                else:
                    drawingMap[y][x] = dummyToIntMap[str(section.data[j][i][k])] + 3000
            if j != (section.dataShape[0] -1):
                if section.column_distance_list[j]:
                    idx[1] = idx[1]+WIDE_DISTANCE_WIDTH
                else:
                    idx[1] = idx[1]+NARROW_DISTANCE_WIDTH
        idx[1] = offset[1]
        if i != (section.dataShape[1] -1):
            if section.row_distance_list[i]:
                idx[0] = idx[0] + WIDE_DISTANCE_WIDTH
            else:
                idx[0] = idx[0] + NARROW_DISTANCE_WIDTH
    return drawingMap

def drawSections(warehouse,s):
    if warehouse.layout == "I":
        idx = [0,0]
        s = drawSection(warehouse.sectionList[0],idx,s)
        idx = [0,warehouse.sectionList[0].dataShape[0]+warehouse.sectionList[0].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH + warehouse.sectionList[0].column_distance_list.count(False) * NARROW_DISTANCE_WIDTH+ROAD_CELL_WIDTH]
        s = drawSection(warehouse.sectionList[1],idx,s)
    
    if warehouse.layout == "H":
        idx = [0,0]
        s = drawSection(warehouse.sectionList[0],idx,s)
        
        idx = [0,np.shape(s)[1] - warehouse.sectionList[3].dataShape[0] - warehouse.sectionList[3].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH - warehouse.sectionList[3].column_distance_list.count(False) * NARROW_DISTANCE_WIDTH]
        s = drawSection(warehouse.sectionList[3],idx,s)
        
        idx = [0,warehouse.sectionList[0].dataShape[0]+warehouse.sectionList[0].column_distance_list.count(True)*WIDE_DISTANCE_WIDTH + +warehouse.sectionList[0].column_distance_list.count(False)*NARROW_DISTANCE_WIDTH+ROAD_CELL_WIDTH]
        s = drawSection(warehouse.sectionList[1],idx,s)
        
        idx = [np.shape(s)[0] - (warehouse.sectionList[2].dataShape[1] * warehouse.sectionList[2].dataShape[2]) - warehouse.sectionList[2].row_distance_list.count(True) * WIDE_DISTANCE_WIDTH - warehouse.sectionList[2].row_distance_list.count(False) * NARROW_DISTANCE_WIDTH,warehouse.sectionList[0].dataShape[0]+warehouse.sectionList[0].column_distance_list.count(True) * WIDE_DISTANCE_WIDTH+warehouse.sectionList[0].column_distance_list.count(False) * NARROW_DISTANCE_WIDTH+ROAD_CELL_WIDTH]
        s = drawSection(warehouse.sectionList[2],idx,s)
    return s

def horizontal_switch(arra):
    a = []
    for i in range(0,len(arra)):
        a.append(arra[len(arra) - i-1])
    return a

def transform(arra):
    _ = np.shape(arra)
    a = []
    for i in range(0,_[1]):
        a.append([])
        for j in range(0,_[0]):
            a[i].append([])
    for i in range(0,_[0]):
        for j in range(0,_[1]):
            a[j][i] = arra[i][j]
    return a
    
def vertical_switch(arra):
    a = []
    for i in range(0,len(arra)):
        arra[i].reverse()
        a.append(arra[i])
    return a

import json
import numpy as np
import random

with open("data.json", "r") as st_json:

    st_python = json.load(st_json)

dummyToIntMap = {}
dummyToColorMap = {}
intToDummyMap = {}
warehousebuf = []
cellCnt = st_python['information']['number_of_warehouse']
for i in range(0,cellCnt):
    sectionList = []
    idx = 0
    sections = st_python['warehouse'][i]['sections']
    sectionCnt = len(st_python['warehouse'][i]['sections'])
    for j in range(0,sectionCnt):
        sectionRow = [(True if p == 0.9 else False )for p in sections[j]['row_distance_list']]
        sectionCol = [(True if p == 0.9 else False )for p in sections[j]['column_distance_list']]
        sectionData = np.shape(st_python['cell_matrix'][i][j])
        data = st_python['cell_matrix'][i][j]
        sectionList.append(section(sectionRow,sectionCol,sectionData,data))
    
    
    sectionList[0].data = vertical_switch(sectionList[0].data)
    
    if st_python['warehouse'][i]['layout'] == "H":
        sectionList[1].data = transform(sectionList[1].data)
        sectionList[1].dataShape = np.shape(sectionList[1].data)
        
        sectionList[2].data = transform(sectionList[2].data)
        sectionList[2].data = vertical_switch(sectionList[2].data)
        sectionList[2].data = horizontal_switch(sectionList[2].data)
        sectionList[2].dataShape = np.shape(sectionList[2].data)
        
        sectionList[3].data = horizontal_switch(sectionList[3].data)
    else:
        sectionList[1].data = horizontal_switch(sectionList[1].data)

    warehousebuf.append(warehouse(st_python['warehouse'][i]['layout'],sectionList))
    
    
    
dummyToIntMap = createDummyIntMap(warehousebuf)
dummyToColorMap = dummyToColor(dummyToIntMap)

intToDummyMap = createIntdymmyMap(dummyToIntMap)

import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Border, Side ,PatternFill

BORDER_THICK = 'thick'
SIDE = Side(border_style=BORDER_THICK,color='000000')

boundDecoder = {"1":Border(left= SIDE,right=SIDE,top=SIDE,),"3":Border(left=SIDE,right=SIDE),"2":Border(left=SIDE,right=SIDE,bottom=SIDE)}
wb = openpyxl.load_workbook(SEEDFILEPATH)


cnt = 1

for _warehouse in warehousebuf:
    ws = wb.create_sheet()
    ws.title = str(cnt)
    cnt = cnt + 1

    ma = getLargeMap(_warehouse)
    a = drawSections(_warehouse,ma)

    for r in a:
        ws.append(list(r))

        
    for colu in range(1,np.shape(a)[1]+1):
        for rows in range(1, np.shape(a)[0]+1):
            value = ws.cell(row = rows, column = colu).value
            border = None
            if int(int(value) / 1000) > 0:
                border = boundDecoder[str(int(value / 1000))]
            value = value % 1000
            color = "ffffff"
            if value == 0:
                pass
            else:
                color = dummyToColorMap[str(value)]
            if value != 0:
                ws.cell(row = rows, column = colu).value = intToDummyMap[int(value)]
            else:
                ws.cell(row = rows, column = colu).value = ' '
            if border != None:
                ws.cell(row = rows, column = colu).border = border
            ws.cell(row = rows, column = colu).fill = PatternFill("solid", start_color=(color))
wb.save(OUTPUTFILEPATH)

